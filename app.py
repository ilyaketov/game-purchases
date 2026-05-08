"""
Веб-приложение Game Purchases — автоматизация отчётов закупа для QuickBooks.

Запуск:
    streamlit run app.py

Дизайн вдохновлён rokky.com:
- Фирменный акцент: индиго #4B4BFF
- Белый фон, минимализм, крупные числа
- Чистая типографика (Inter), большие отступы
- Прогрессивное раскрытие шагов (1 → 2 → 3)
"""
from __future__ import annotations

import hashlib
import io
import shutil
import tempfile
import time
import zipfile
from pathlib import Path

import pandas as pd
import streamlit as st

from engine import Pipeline
from config import PLOSHADKA_MAP


# ===========================================================================
# Helper для безопасной вставки HTML (не парсится как markdown)
# ===========================================================================
def _html(html: str) -> None:
    """Вставка HTML без markdown-парсинга. Главное преимущество перед
    st.markdown(unsafe_allow_html=True) — отступы в многострочных f-strings
    НЕ интерпретируются как indented code block.
    """
    st.html(html)



# ===========================================================================
# Авто-определение типа загруженного файла
# ===========================================================================
KIND_LABEL = {
    "r1":    "Универсальный отчёт",
    "r2":    "Universal Report shipped",
    "genba": "genbaFile",
}


def _detect_file_kind(path: Path) -> str | None:
    """Определяет, что за файл, по заголовкам первой строки первого листа.

    Возвращает 'r1' / 'r2' / 'genba' / None.
    Чтение через openpyxl read_only — занимает 50мс…5с, не зависит от размера.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header = next(ws.iter_rows(max_row=1, values_only=True))
        wb.close()
        cols = set(c for c in header if c is not None)
    except Exception:
        return None

    # genbaFile — самые специфичные колонки
    if "Activation Qty" in cols and "Grand Total" in cols:
        return "genba"
    # R1 — старый формат, ВАЖНО: "площадка " с пробелом
    if "площадка " in cols:
        return "r1"
    # R2 — новый shipped, "площадка" без пробела + есть Количество
    if "площадка" in cols and "Количество" in cols:
        return "r2"
    return None



# ===========================================================================
# Helpers для инфографики
# ===========================================================================
# Палитра — оттенки фирменного индиго от насыщенного к светлому
PALETTE = ["#4B4BFF", "#7C5CFF", "#9F8DFF", "#B6A8FF", "#CFC5FF",
                  "#DAD3FF", "#E8E3FF", "#EFEAFF", "#F5F2FF"]


def _fmt_money(v: float) -> str:
    """Форматирует $1234567.89 → '$1 234 567.89'."""
    return f"${v:,.2f}".replace(",", " ")


def _fmt_int(v: int) -> str:
    """Форматирует 1234567 → '1 234 567'."""
    return f"{v:,}".replace(",", " ")


def _render_breakdown_list(success: dict) -> None:
    """Концепт 4 — компактный список со всеми площадками + прогресс-бары."""
    items = sorted(success.items(), key=lambda x: -x[1]["cost"])
    total_cost = sum(v["cost"] for _, v in items)
    total_qty = sum(v["qty"] for _, v in items)

    rows_html = []
    for idx, (name, data) in enumerate(items):
        color = PALETTE[min(idx, len(PALETTE) - 1)]
        pct = data["cost"] / total_cost * 100 if total_cost else 0
        avg = data["cost"] / data["qty"] if data["qty"] else 0
        rows_html.append(f"""
        <div style="display: grid; grid-template-columns: 24px 110px 90px 1fr 130px 90px;
                    gap: 0.75rem; align-items: center; padding: 0.85rem 1.25rem;
                    border-bottom: 1px solid #F4F4F8;">
          <span style="display: inline-block; width: 8px; height: 8px; border-radius: 2px; background: {color};"></span>
          <span style="font-weight: 600; color: #0A0A1F; font-size: 0.9rem;">{name}</span>
          <span style="font-variant-numeric: tabular-nums; text-align: right; color: #0A0A1F; font-weight: 500; font-size: 0.85rem;">{_fmt_int(data['qty'])}</span>
          <div style="display: flex; align-items: center; gap: 0.5rem;">
            <div style="flex: 1; height: 6px; background: #F0F0F4; border-radius: 99px; overflow: hidden;">
              <div style="width: {pct:.1f}%; height: 100%; background: {color}; border-radius: 99px;"></div>
            </div>
            <span style="font-size: 0.75rem; color: #6B6B80; min-width: 40px; text-align: right; font-variant-numeric: tabular-nums;">{pct:.1f}%</span>
          </div>
          <span style="font-variant-numeric: tabular-nums; text-align: right; color: #0A0A1F; font-weight: 500; font-size: 0.85rem;">{_fmt_money(data['cost'])}</span>
          <span style="font-variant-numeric: tabular-nums; text-align: right; color: #6B6B80; font-size: 0.8rem;">{_fmt_money(avg)}</span>
        </div>""")

    avg_total = total_cost / total_qty if total_qty else 0
    total_row_html = f"""
    <div style="display: grid; grid-template-columns: 24px 110px 90px 1fr 130px 90px;
                gap: 0.75rem; align-items: center; padding: 0.85rem 1.25rem;
                background: #FAFAFB; border-top: 1px solid #EDEDF2;">
      <span></span>
      <span style="font-weight: 700; color: #0A0A1F; font-size: 0.9rem;">Итого</span>
      <span style="font-variant-numeric: tabular-nums; text-align: right; color: #0A0A1F; font-weight: 700; font-size: 0.9rem;">{_fmt_int(total_qty)}</span>
      <span></span>
      <span style="font-variant-numeric: tabular-nums; text-align: right; color: #0A0A1F; font-weight: 700; font-size: 0.9rem;">{_fmt_money(total_cost)}</span>
      <span style="font-variant-numeric: tabular-nums; text-align: right; color: #6B6B80; font-size: 0.8rem;">{_fmt_money(avg_total)}</span>
    </div>
    """

    header_html = """
    <div style="display: grid; grid-template-columns: 24px 110px 90px 1fr 130px 90px;
                gap: 0.75rem; align-items: center; padding: 0.6rem 1.25rem 0.5rem;
                font-size: 0.7rem; color: #8B8B9E; text-transform: uppercase;
                letter-spacing: 0.04em; font-weight: 500; border-bottom: 1px solid #EDEDF2;">
      <span></span>
      <span>Площадка</span>
      <span style="text-align: right;">Ключи</span>
      <span style="padding-left: 0.5rem;">Доля от всего</span>
      <span style="text-align: right;">Себестоимость</span>
      <span style="text-align: right;">Avg / ключ</span>
    </div>
    """

    list_html = (
        '<div style="background: white; border: 1px solid #EDEDF2; border-radius: 14px; overflow: hidden;">'
        + header_html + ''.join(rows_html) + total_row_html + '</div>'
    )
    _html(list_html)


# ===========================================================================
# Настройка страницы
# ===========================================================================
st.set_page_config(
    page_title="Game Purchases",
    page_icon="◆",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# ===========================================================================
# Стили — Rokky-inspired
# ===========================================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

/* === Глобальные === */
.stApp {
    background: #FAFAFB;
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}

/* НЕ ставим font-family глобально на все элементы — иначе ломаются иконки-лигатуры
   (Streamlit использует Material Icons со словами-лигатурами вроде "upload", "menu",
   которые без правильного шрифта рендерятся как обычный текст и накладываются на labels). */

/* Сужаем основной контейнер */
.main .block-container {
    max-width: 1100px;
    padding-top: 2.5rem;
    padding-bottom: 4rem;
}

/* === Прячем стандартный header === */
header[data-testid="stHeader"] {
    background: transparent;
    height: 0;
}

#MainMenu, footer {
    visibility: hidden;
}

/* === Заголовки === */
h1 {
    font-weight: 800 !important;
    font-size: 3rem !important;
    letter-spacing: -0.03em !important;
    color: #0A0A1F !important;
    line-height: 1.1 !important;
    margin-bottom: 0.5rem !important;
}

h2 {
    font-weight: 700 !important;
    font-size: 1.5rem !important;
    color: #0A0A1F !important;
    letter-spacing: -0.01em !important;
    margin-top: 2.5rem !important;
    margin-bottom: 1.25rem !important;
}

h3 {
    font-weight: 600 !important;
    font-size: 1rem !important;
    color: #0A0A1F !important;
    margin-bottom: 0.75rem !important;
}

p, .stMarkdown p {
    color: #4A4A5E;
    font-size: 0.95rem;
    line-height: 1.6;
}

/* Hero subtitle */
.hero-subtitle {
    font-size: 1.15rem;
    color: #6B6B80;
    font-weight: 400;
    margin-top: 0.25rem;
    margin-bottom: 2.5rem;
    max-width: 640px;
}

/* Step number */
.step-num {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 28px;
    height: 28px;
    background: #4B4BFF;
    color: white;
    border-radius: 50%;
    font-size: 0.85rem;
    font-weight: 600;
    margin-right: 0.6rem;
    vertical-align: middle;
}

.step-row {
    display: flex;
    align-items: center;
    margin-top: 2.5rem;
    margin-bottom: 1.25rem;
}

.step-row h2 {
    margin: 0 !important;
}

/* === Кнопки === */
.stButton > button, .stDownloadButton > button {
    font-weight: 500 !important;
    font-size: 0.9rem !important;
    border-radius: 8px !important;
    padding: 0.6rem 1.4rem !important;
    transition: all 0.15s ease !important;
    border: 1px solid transparent !important;
}

/* Primary кнопка — фирменный индиго.
   Используем широкий селектор, чтобы покрыть и старые (kind="primary"),
   и новые (data-testid="stBaseButton-primary") версии Streamlit. */
.stButton > button[kind="primary"],
.stDownloadButton > button[kind="primary"],
button[data-testid="stBaseButton-primary"],
button[data-testid="stBaseButton-primaryFormSubmit"] {
    background: #4B4BFF !important;
    color: #FFFFFF !important;
    border: none !important;
    box-shadow: 0 1px 2px rgba(75, 75, 255, 0.15) !important;
}

/* Текст внутри primary кнопки — гарантируем белый */
.stButton > button[kind="primary"] *,
.stDownloadButton > button[kind="primary"] *,
button[data-testid="stBaseButton-primary"] *,
button[data-testid="stBaseButton-primaryFormSubmit"] * {
    color: #FFFFFF !important;
}

.stButton > button[kind="primary"]:hover,
.stDownloadButton > button[kind="primary"]:hover,
button[data-testid="stBaseButton-primary"]:hover,
button[data-testid="stBaseButton-primaryFormSubmit"]:hover {
    background: #3838E5 !important;
    color: #FFFFFF !important;
    transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(75, 75, 255, 0.25) !important;
}

/* Secondary кнопка */
.stButton > button[kind="secondary"],
.stDownloadButton > button[kind="secondary"],
button[data-testid="stBaseButton-secondary"] {
    background: white !important;
    color: #0A0A1F !important;
    border: 1px solid #E4E4EA !important;
}

.stButton > button[kind="secondary"]:hover,
.stDownloadButton > button[kind="secondary"]:hover,
button[data-testid="stBaseButton-secondary"]:hover {
    border-color: #4B4BFF !important;
    color: #4B4BFF !important;
}

/* === Загрузка файлов — оставляем нативный вид Streamlit, кастомизируем только обёртку === */
[data-testid="stFileUploader"] > section {
    border-radius: 14px !important;
}

/* === Метрики (Streamlit metric) === */
[data-testid="stMetric"] {
    background: white;
    padding: 1.25rem 1.5rem;
    border-radius: 12px;
    border: 1px solid #EDEDF2;
}

[data-testid="stMetric"] [data-testid="stMetricLabel"] {
    color: #6B6B80;
    font-size: 0.75rem !important;
    font-weight: 500 !important;
    text-transform: uppercase;
    letter-spacing: 0.04em;
}

[data-testid="stMetric"] [data-testid="stMetricValue"] {
    color: #0A0A1F;
    font-size: 1.75rem !important;
    font-weight: 700 !important;
    letter-spacing: -0.02em;
}

/* === Карточки-обёртки для блоков === */
.kf-card {
    background: white;
    border: 1px solid #EDEDF2;
    border-radius: 14px;
    padding: 1.5rem 1.75rem;
    margin-bottom: 1rem;
}

.kf-card-uploader {
    background: white;
    border: 1px solid #EDEDF2;
    border-radius: 14px;
    padding: 1.25rem;
    height: 100%;
}

.kf-uploader-label {
    font-weight: 600;
    font-size: 0.95rem;
    color: #0A0A1F;
    margin-bottom: 0.25rem;
}

.kf-uploader-hint {
    font-size: 0.8rem;
    color: #8B8B9E;
    margin-bottom: 1rem;
}

/* === Алёрты === */
[data-testid="stAlert"] {
    border-radius: 10px;
    border: none;
    padding: 0.85rem 1.1rem;
    font-size: 0.9rem;
}

div[data-baseweb="notification"][kind="info"] {
    background: #EEF0FF;
}

/* === Multiselect === */
[data-baseweb="select"] {
    border-radius: 10px !important;
}

[data-baseweb="tag"] {
    background: #EEF0FF !important;
    color: #4B4BFF !important;
    border-radius: 6px !important;
    font-weight: 500 !important;
}

/* === Таблицы === */
[data-testid="stDataFrame"] {
    border-radius: 12px;
    border: 1px solid #EDEDF2;
    overflow: hidden;
}

/* === Прогресс === */
.stProgress > div > div > div > div {
    background: #4B4BFF !important;
}

/* === Expander === */
[data-testid="stExpander"] {
    background: white;
    border: 1px solid #EDEDF2 !important;
    border-radius: 10px !important;
}

[data-testid="stExpander"] summary {
    font-weight: 500;
    font-size: 0.9rem;
}

/* === Caption === */
.caption-muted {
    color: #8B8B9E;
    font-size: 0.85rem;
    margin-top: 0.5rem;
}

/* Status pill */
.kf-pill {
    display: inline-block;
    padding: 0.2rem 0.7rem;
    background: #E8F5EE;
    color: #0F7A3E;
    font-size: 0.75rem;
    font-weight: 500;
    border-radius: 999px;
    letter-spacing: 0.01em;
}

.kf-pill-warn {
    background: #FFF4E5;
    color: #B25E00;
}

/* Логотип */
.kf-brand {
    display: flex;
    align-items: center;
    gap: 0.6rem;
    margin-bottom: 1rem;
}

.kf-brand-mark {
    width: 32px;
    height: 32px;
    background: linear-gradient(135deg, #4B4BFF 0%, #7C5CFF 100%);
    border-radius: 8px;
    display: flex;
    align-items: center;
    justify-content: center;
    color: white;
    font-weight: 700;
    font-size: 1rem;
    box-shadow: 0 2px 8px rgba(75, 75, 255, 0.25);
}

.kf-brand-name {
    font-size: 0.95rem;
    font-weight: 600;
    color: #0A0A1F;
    letter-spacing: -0.01em;
}

.kf-brand-sub {
    font-size: 0.85rem;
    color: #8B8B9E;
}

/* Discrete вертикальный отступ */
.kf-spacer {
    height: 1.5rem;
}

</style>
""", unsafe_allow_html=True)


# ===========================================================================
# Шапка / Hero
# ===========================================================================
st.markdown("""
<div class="kf-brand">
  <div class="kf-brand-mark">◆</div>
  <div>
    <div class="kf-brand-name">Game Purchases</div>
    <div class="kf-brand-sub">Purchase reports automation</div>
  </div>
</div>
""", unsafe_allow_html=True)

st.markdown("# Закупка игр", unsafe_allow_html=True)
st.markdown(
    '<p class="hero-subtitle">Загрузите выгрузки из биллинга — '
    'приложение соберёт сводные отчёты для всех площадок в формате QuickBooks. '
    'Минуты вместо часов ручной работы.</p>',
    unsafe_allow_html=True,
)


# ===========================================================================
# Временная директория и кэш
# ===========================================================================
@st.cache_resource
def _make_tmpdir() -> Path:
    return Path(tempfile.mkdtemp(prefix="keyflow_"))


tmpdir = _make_tmpdir()


@st.cache_resource(show_spinner=False)
def _build_pipeline(r1_path: str | None, r2_path: str | None,
                    genba_path: str | None, _cache_key: str) -> Pipeline:
    return Pipeline(r1_path, r2_path, genba_path)


# ===========================================================================
# Шаг 1 — загрузка
# ===========================================================================
st.markdown(
    '<div class="step-row"><span class="step-num">1</span>'
    '<h2>Загрузите отчёты</h2></div>',
    unsafe_allow_html=True,
)

st.caption(
    "Перетащите все три файла сразу — Универсальный отчёт, Universal Report shipped "
    "и genbaFile. Тип определится автоматически по заголовкам столбцов."
)

uploaded = st.file_uploader(
    "files",
    type=["xlsx"],
    accept_multiple_files=True,
    label_visibility="collapsed",
)

if not uploaded:
    st.markdown('<div class="kf-spacer"></div>', unsafe_allow_html=True)
    st.info("Загрузите хотя бы один файл, чтобы продолжить.")
    st.stop()

# -- Сохраняем все загруженные файлы и определяем их тип -----------------------
def _save_and_detect(uploaded_file):
    """Сохраняет файл во tmpdir, возвращает (path, sha1, kind, size_bytes)."""
    data = uploaded_file.getbuffer()
    digest = hashlib.sha1(bytes(data)).hexdigest()[:16]
    target = tmpdir / f"{digest}_{uploaded_file.name}"
    if not target.exists():
        target.write_bytes(data)
    return target, digest, _detect_file_kind(target), len(data)


with st.spinner("Распознаём файлы..."):
    detected = []  # [(orig_name, path, sha1, kind, size)]
    for f in uploaded:
        path, sha1, kind, size = _save_and_detect(f)
        detected.append((f.name, path, sha1, kind, size))

# Группируем по kind. Если несколько файлов одного типа — берём последний.
by_kind: dict[str, dict] = {}
unknown_files: list[str] = []
for orig_name, path, sha1, kind, size in detected:
    if kind is None:
        unknown_files.append(orig_name)
        continue
    by_kind[kind] = {
        "name": orig_name, "path": path, "sha1": sha1, "size": size,
    }

# -- Карточки с распознанными файлами ----------------------------------------
def _fmt_size(b: int) -> str:
    if b > 1024 * 1024:
        return f"{b / 1024 / 1024:.1f} МБ"
    return f"{b / 1024:.0f} КБ"


cards_html = []
for kind in ["r1", "r2", "genba"]:
    if kind in by_kind:
        info = by_kind[kind]
        cards_html.append(f"""
        <div style="background: white; border: 1px solid #EDEDF2; border-radius: 12px;
                    padding: 0.85rem 1rem; display: flex; align-items: center; gap: 0.75rem;">
          <div style="width: 28px; height: 28px; border-radius: 50%; background: #E8F5EE;
                      color: #0F7A3E; display: flex; align-items: center; justify-content: center;
                      font-size: 1rem; flex-shrink: 0;">✓</div>
          <div style="min-width: 0; flex: 1;">
            <div style="font-weight: 600; font-size: 0.85rem; color: #0A0A1F;">{KIND_LABEL[kind]}</div>
            <div style="font-size: 0.75rem; color: #8B8B9E; white-space: nowrap;
                        overflow: hidden; text-overflow: ellipsis;">
              {info['name']} · {_fmt_size(info['size'])}
            </div>
          </div>
        </div>
        """)
    else:
        cards_html.append(f"""
        <div style="background: white; border: 1px dashed #E4E4EA; border-radius: 12px;
                    padding: 0.85rem 1rem; display: flex; align-items: center; gap: 0.75rem;">
          <div style="width: 28px; height: 28px; border-radius: 50%; background: #F4F4F8;
                      color: #B0B0C0; display: flex; align-items: center; justify-content: center;
                      font-size: 0.95rem; flex-shrink: 0;">○</div>
          <div style="min-width: 0; flex: 1;">
            <div style="font-weight: 500; font-size: 0.85rem; color: #6B6B80;">{KIND_LABEL[kind]}</div>
            <div style="font-size: 0.75rem; color: #B0B0C0;">не загружен</div>
          </div>
        </div>
        """)

_html(
    f'<div style="display: grid; grid-template-columns: repeat(3, 1fr); '
    f'gap: 0.75rem; margin-top: 1rem;">{"".join(cards_html)}</div>'
)

if unknown_files:
    st.markdown('<div class="kf-spacer"></div>', unsafe_allow_html=True)
    st.warning(
        "Не удалось распознать: " + ", ".join(f"`{n}`" for n in unknown_files) +
        ". Похоже, это не Универсальный отчёт и не genbaFile — проверьте, "
        "что файл выгружен из биллинга в нужном формате."
    )

# Нужен хотя бы один R1 или R2 для продолжения
if not (by_kind.get("r1") or by_kind.get("r2")):
    st.markdown('<div class="kf-spacer"></div>', unsafe_allow_html=True)
    st.info("Загрузите хотя бы один из универсальных отчётов, чтобы продолжить.")
    st.stop()

# Готовим параметры для pipeline
p1_path = by_kind.get("r1", {}).get("path")
p2_path = by_kind.get("r2", {}).get("path")
pg_path = by_kind.get("genba", {}).get("path")
h1 = by_kind.get("r1", {}).get("sha1")
h2 = by_kind.get("r2", {}).get("sha1")
hg = by_kind.get("genba", {}).get("sha1")
cache_key = f"{h1}|{h2}|{hg}"


# ===========================================================================
# Шаг 2 — анализ
# ===========================================================================
st.markdown(
    '<div class="step-row"><span class="step-num">2</span>'
    '<h2>Анализ</h2></div>',
    unsafe_allow_html=True,
)

with st.spinner("Парсим файлы (30–60 секунд при первой загрузке, дальше — мгновенно)..."):
    t0 = time.time()
    pipeline = _build_pipeline(
        str(p1_path) if p1_path else None,
        str(p2_path) if p2_path else None,
        str(pg_path) if pg_path else None,
        cache_key,
    )
    load_time = time.time() - t0

with st.spinner("Проверяем поставщиков..."):
    validation = pipeline.validate()

if not validation.available_ploshadki:
    st.error("В загруженных файлах не найдено ни одной известной площадки.")
    st.stop()

# Pill со статусом
n_ok = "Все распознаны" if not validation.unmapped_suppliers else f"{len(validation.unmapped_suppliers)} требует внимания"
pill_class = "kf-pill" if not validation.unmapped_suppliers else "kf-pill kf-pill-warn"
st.markdown(
    f'<span class="{pill_class}">{n_ok}</span> '
    f'<span class="caption-muted">·  загрузка за {load_time:.1f} с  ·  '
    f'{len(validation.available_ploshadki)} площадок найдено</span>',
    unsafe_allow_html=True,
)

st.markdown('<div class="kf-spacer"></div>', unsafe_allow_html=True)

# Метрики площадок — компактная сетка
n_ploshadki = len(validation.available_ploshadki)
n_per_row = min(n_ploshadki, 5)
items = list(validation.available_ploshadki.items())

for i in range(0, n_ploshadki, n_per_row):
    cols = st.columns(n_per_row, gap="small")
    chunk = items[i:i + n_per_row]
    for col, (key, n) in zip(cols, chunk):
        with col:
            st.metric(key, f"{n:,}".replace(",", " "))

# Неизвестные поставщики
if validation.unmapped_suppliers:
    st.markdown('<div class="kf-spacer"></div>', unsafe_allow_html=True)
    with st.expander(
        f"Неизвестные поставщики ({len(validation.unmapped_suppliers)}) — будут пропущены",
        expanded=False,
    ):
        st.caption(
            "Эти имена не сопоставлены с группами в эталоне. "
            "Чтобы строки попали в отчёт, добавьте маппинги в `SUPPLIER_MAPPING` (config.py)."
        )
        df_unmapped = pd.DataFrame([
            {"Имя в биллинге": k, "Строк": v}
            for k, v in sorted(validation.unmapped_suppliers.items(),
                               key=lambda x: -x[1])
        ])
        st.dataframe(df_unmapped, use_container_width=True, hide_index=True)


# ===========================================================================
# Шаг 3 — сборка
# ===========================================================================
st.markdown(
    '<div class="step-row"><span class="step-num">3</span>'
    '<h2>Соберите отчёты</h2></div>',
    unsafe_allow_html=True,
)

selected = st.multiselect(
    "Площадки",
    options=list(validation.available_ploshadki.keys()),
    default=list(validation.available_ploshadki.keys()),
    label_visibility="collapsed",
    placeholder="Выберите площадки",
)

if not selected:
    st.info("Выберите хотя бы одну площадку.")
    st.stop()

st.markdown('<div class="kf-spacer"></div>', unsafe_allow_html=True)

run_clicked = st.button("Собрать отчёты", type="primary", use_container_width=False)

if run_clicked:
    out_dir = tmpdir / f"out_{cache_key[:16]}"
    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir()

    progress = st.progress(0, text="Запуск...")
    results = {}
    for i, key in enumerate(selected):
        progress.progress(i / len(selected), text=f"{key}...")
        try:
            t0 = time.time()
            agg = pipeline.aggregate(key)
            if agg.empty:
                results[key] = {"error": "Нет данных"}
                continue
            path = pipeline.save_to_excel(
                agg, key, out_dir / f"{key}_zakup_svod.xlsx"
            )
            active = agg[agg["qty"] > 0]
            results[key] = {
                "path": path,
                "qty": int(active["qty"].sum()),
                "cost": float(active["cost"].sum()),
                "suppliers": int(active["supplier_group"].nunique()),
                "products": int(len(active)),
                "elapsed": time.time() - t0,
            }
        except Exception as e:
            results[key] = {"error": str(e)}

    progress.empty()
    # Сохраняем результаты в session_state, чтобы пережить перерисовку при кликах
    st.session_state["results"] = results
    st.session_state["out_dir"] = str(out_dir)

# Если результаты собраны (или ранее, или только что) — показываем
if "results" in st.session_state:
    results = st.session_state["results"]
    success = {k: v for k, v in results.items() if "error" not in v}
    failed = {k: v for k, v in results.items() if "error" in v}

    if success:
        # ===================================================================
        # Компактный список со всеми площадками
        # ===================================================================
        st.markdown(
            '<h3 style="margin-top: 1.5rem; margin-bottom: 1rem;">Готово</h3>',
            unsafe_allow_html=True,
        )
        _render_breakdown_list(success)

        st.markdown('<div class="kf-spacer"></div>', unsafe_allow_html=True)

        # ===================================================================
        # Скачивание
        # ===================================================================
        st.markdown(
            '<h3 style="margin-top: 2rem; margin-bottom: 0.75rem;">Скачать</h3>',
            unsafe_allow_html=True,
        )

        # ZIP со всеми
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for k, v in success.items():
                zf.write(v["path"], arcname=Path(v["path"]).name)
        zip_buffer.seek(0)

        dl_cols = st.columns([1, 3])
        with dl_cols[0]:
            st.download_button(
                "Скачать всё (ZIP)",
                data=zip_buffer,
                file_name="keyflow_reports.zip",
                mime="application/zip",
                type="primary",
                use_container_width=True,
            )

        st.markdown('<div class="kf-spacer"></div>', unsafe_allow_html=True)
        st.markdown(
            '<p style="font-size: 0.85rem; color: #8B8B9E; margin-bottom: 0.5rem;">'
            'Или по отдельности:</p>',
            unsafe_allow_html=True,
        )

        files_per_row = 4
        items_l = list(success.items())
        for i in range(0, len(items_l), files_per_row):
            row_cols = st.columns(files_per_row, gap="small")
            chunk = items_l[i:i + files_per_row]
            for col, (k, v) in zip(row_cols, chunk):
                with col:
                    with open(v["path"], "rb") as f:
                        st.download_button(
                            k,
                            data=f.read(),
                            file_name=Path(v["path"]).name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_{k}",
                            use_container_width=True,
                        )

    if failed:
        st.markdown('<div class="kf-spacer"></div>', unsafe_allow_html=True)
        st.error("Не удалось обработать:")
        for k, v in failed.items():
            st.write(f"**{k}** — {v['error']}")


# ===========================================================================
# Подвал
# ===========================================================================
st.markdown('<div style="height: 4rem;"></div>', unsafe_allow_html=True)
st.markdown(
    '<div style="text-align: center; color: #B0B0C0; font-size: 0.8rem;">'
    'Game Purchases · Purchase reports automation</div>',
    unsafe_allow_html=True,
)
